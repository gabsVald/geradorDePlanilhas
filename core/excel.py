"""
================================================================================
core/excel.py — Geração de Planilhas Excel (.xlsm)
================================================================================
Responsável por montar e gravar o arquivo Excel de corte a partir dos blocos
de itens processados. Opera sobre um molde .xlsm com macros VBA preservadas.

Fluxo principal:
  1. Carrega o molde e ajusta o número de linhas dinamicamente
  2. Escreve cabeçalho (logo/marca, título, data, quantidade)
  3. Itera sobre os blocos (normal e prensado), escrevendo cada item
  4. Aplica regras especiais: fitas de borda, veio, +5mm/+25mm, FORMICA, VIROLA FLEX
  5. Salva via arquivo temporário para evitar corrupção em falhas de disco/rede
================================================================================
"""

# ===== IMPORTAÇÕES =====
import os
import re
from pathlib import Path
from datetime import datetime
from copy import copy  # Usado para copiar estilos de célula do molde

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage  # Inserção de logos PNG
from openpyxl.styles import Font, Alignment, PatternFill   # Formatação de células

from utils.config import REGRAS           # Dicionário global de regras (regras.json)
from utils.helpers import (
    limpar,                    # Normaliza strings vindas do PDM/Excel
    converter_para_numero,     # Converte string para int arredondado
    limpar_material_rigoroso,  # Remove sujeiras do nome do material
    resource_path              # Resolve caminhos dentro do .exe (PyInstaller)
)


# ===== FUNÇÕES AUXILIARES =====

def buscar_valor_valido(item, indices):
    """
    Percorre uma lista de índices e retorna o primeiro valor numérico válido (> 0).

    Utilizada para buscar comprimento, largura e espessura com fallback entre
    colunas do PDM, pois o layout pode variar conforme o tipo de item.

    Parâmetros:
        item (dict): Dicionário de dados da peça (chaves = índices de coluna).
        indices (list[int]): Ordem de prioridade das colunas a tentar.

    Retorna:
        int | float: Primeiro valor numérico > 0 encontrado, ou 0 se nenhum servir.
    """
    for idx in indices:
        val = item.get(idx)
        if val is not None:
            texto = str(val).strip().lower()
            # Ignora células vazias, NaN, None ou zero explícito
            if texto not in ["", "nan", "none", "0", "0,0", "0.0"]:
                num = converter_para_numero(val)
                if num and num > 0:
                    return num
    return 0


def escrever_seguro(ws, coord, valor, alinhamento=None):
    """
    Escreve um valor em uma célula, tratando o caso de célula mesclada.

    O openpyxl não permite escrever diretamente em células mescladas secundárias;
    é necessário localizar a célula-mestre do intervalo de mesclagem.

    Parâmetros:
        ws: Worksheet ativa do openpyxl.
        coord (str): Coordenada da célula (ex: 'B3', 'A10').
        valor: Valor a escrever.
        alinhamento (Alignment | None): Alinhamento opcional a aplicar.
    """
    try:
        cell = ws[coord]
        if cell.__class__.__name__ == 'MergedCell':
            # Célula mesclada: encontra a mestre (canto superior esquerdo)
            for r in ws.merged_cells.ranges:
                if coord in r:
                    m = ws.cell(row=r.min_row, column=r.min_col)
                    m.value = valor
                    if alinhamento:
                        m.alignment = alinhamento
                    return
        else:
            cell.value = valor
            if alinhamento:
                cell.alignment = alinhamento
    except Exception as e:
        print(f"[EXCEL WARN] Falha ao escrever na célula {coord}: {e}")


def tratar_cabecalho_a1(ws, id_projeto):
    """
    Preenche a célula A1 com o logo ou nome da marca do projeto.

    Prioridade:
      1. Marcas configuradas como TEXTO (ex: ZARA, ZAFFARI) — escrito em Arial 22 bold.
      2. Marcas configuradas como IMAGEM — PNG de 100x35px inserido em A1.
      3. Fallback: logo padrão da Ingecon.

    Parâmetros:
        ws: Worksheet ativa.
        id_projeto (str): Código do projeto (ex: 'ZAR001', 'ING123').
    """
    marcas_texto = REGRAS["especiais"]["marcas_texto"]
    logos_imagem = REGRAS["especiais"]["marcas_imagem"]
    id_up = str(id_projeto).upper()
    ws['A1'].value = None  # Limpa conteúdo anterior antes de sobrescrever

    # Verifica se alguma sigla de marca-texto está no código do projeto
    for sigla, nome in marcas_texto.items():
        if sigla in id_up:
            escrever_seguro(ws, 'A1', nome, Alignment(horizontal='center', vertical='center'))
            ws['A1'].font = Font(name='Arial', size=22, bold=True)
            return

    # Verifica se existe um PNG mapeado para a marca
    for sigla, arq in logos_imagem.items():
        if sigla in id_up:
            path = resource_path(f"logos/{arq}.png")
            if Path(path).exists():
                img = OpenpyxlImage(path)
                img.width, img.height = 100, 35  # Dimensões padrão do logo (px)
                ws.row_dimensions[1].height = 33  # Altura da linha 1 em pontos
                ws.add_image(img, 'A1')
                return

    # Fallback: logo padrão da Ingecon se nenhuma marca for reconhecida
    path_ing = resource_path("logos/ingecon.png")
    if Path(path_ing).exists():
        img = OpenpyxlImage(path_ing)
        img.width, img.height = 100, 35
        ws.row_dimensions[1].height = 33
        ws.add_image(img, 'A1')


def ajustar_molde_elastico(ws, num_itens):
    """
    Adapta dinamicamente o molde .xlsm para acomodar o número exato de itens.

    O molde tem 3 linhas padrão de dados. Se houver mais ou menos itens,
    o molde é expandido (insert_rows) ou contraído (delete_rows) conforme necessário.
    As mesclagens do rodapé são desfeitas antes e refeitas após o ajuste.

    Parâmetros:
        ws: Worksheet ativa.
        num_itens (int): Total de linhas de dados a escrever (incluindo cabeçalhos de prensado).

    Retorna:
        int: Número da linha onde começa a área de observações (rodapé).
    """
    # Torna todas as linhas visíveis para evitar conflito com linhas ocultas do molde
    for r in range(1, 50):
        ws.row_dimensions[r].hidden = False

    padrao = 3       # Número de linhas de dados no molde original
    l_rodape = 9     # Linha onde começa o rodapé (área de observações)
    quadro = None    # Armazena info da mesclagem do rodapé para recriar depois

    # Desfaz mesclagens do rodapé para poder inserir/remover linhas sem conflito
    for m in list(ws.merged_cells.ranges):
        if m.min_row >= l_rodape:
            if m.min_row == l_rodape:
                # Salva as colunas do quadro de observações para recriar depois
                quadro = {'min_col': m.min_col, 'max_col': m.max_col, 'max_row': m.max_row}
            try:
                ws.unmerge_cells(str(m))
            except Exception:
                pass

    # Remove linhas extras do quadro se existiam (molde pode ter linhas internas mescladas)
    if quadro and quadro['max_row'] > l_rodape:
        ws.delete_rows(l_rodape + 1, quadro['max_row'] - l_rodape)

    # Insere ou remove linhas para ajustar ao número de itens
    diff = num_itens - padrao
    if diff > 0:
        ws.insert_rows(l_rodape, diff)   # Expande: insere linhas antes do rodapé
    elif diff < 0:
        ws.delete_rows(l_rodape + diff, abs(diff))  # Contrai: remove linhas excedentes

    font_pecas = Font(name='Arial', size=10)  # Fonte padrão para linhas de dados

    # Configura cada linha de dados: altura, colunas X, fonte e cópia de estilo
    for r in range(6, 6 + num_itens):
        ws.row_dimensions[r].height = 25.5  # Altura padrão das linhas de peças
        # Colunas D (4) e G (7) sempre recebem "X" (separadores visuais no molde)
        ws.cell(row=r, column=4).value = ws.cell(row=r, column=7).value = "X"

        for c in range(1, 16):
            ws.cell(row=r, column=c).font = font_pecas
            if r > 6:
                # Copia o estilo da linha 6 (primeira linha de dados) para as demais
                src, tgt = ws.cell(row=6, column=c), ws.cell(row=r, column=c)
                if src.has_style:
                    tgt._style = copy(src._style)

    # Recria a mesclagem do quadro de observações na nova posição
    n_inicio = 6 + num_itens
    if quadro:
        ws.merge_cells(
            start_row=n_inicio, start_column=quadro['min_col'],
            end_row=n_inicio, end_column=quadro['max_col']
        )
    return n_inicio


def _sufixo_dimensional(itens, fonte_espessura):
    """
    Calcula o sufixo dimensional consolidado para blocos prensados.

    Para blocos prensados, a descrição da peça é substituída pelas dimensões
    máximas do conjunto menos 10mm (margem de corte), no formato CxLxE.

    Parâmetros:
        itens (list[dict]): Lista de itens do bloco prensado.
        fonte_espessura (dict): Dicionário do pai ou prensado_info com a espessura.

    Retorna:
        str: String formatada como " 590X290X18", ou "" se alguma dimensão faltar.
    """
    if not itens:
        return ""

    # Comprimento e largura: maior valor entre os itens, menos 10mm de margem
    max_c = max((buscar_valor_valido(i, [8, 15]) for i in itens), default=0)
    max_l = max((buscar_valor_valido(i, [10, 16]) for i in itens), default=0)

    # Espessura: vem sempre do título do bloco prensado, não dos itens internos
    v_e = buscar_valor_valido(fonte_espessura, [12, 13])

    v_c = max_c - 10 if max_c else 0
    v_l = max_l - 10 if max_l else 0

    if v_c > 0 and v_l > 0 and v_e:
        return f" {int(v_c)}X{int(v_l)}X{int(v_e)}"
    return ""  # Retorna vazio se alguma dimensão estiver ausente


# ===== FUNÇÃO PRINCIPAL DE GERAÇÃO =====

def gerar_arquivo_excel(pai, blocos, id_proj, qtd_tot, molde, pasta, pai_is_prensado):
    """
    Gera o arquivo .xlsm de plano de corte a partir dos dados processados.

    Parâmetros:
        pai (Series|dict): Linha do DataFrame com dados do item pai (código, descrição, acabamento).
        blocos (list[dict]): Lista de blocos de itens {'tipo': 'normal'|'prensado', 'itens': [...]}.
        id_proj (str): Código do projeto (ex: 'ZAR00123') — usado para logo e nome de arquivo.
        qtd_tot (float): Quantidade total do pai (vai para célula A3 como multiplicador).
        molde (Path): Caminho do arquivo .xlsm de molde com macros VBA.
        pasta (str): Diretório de destino para salvar o arquivo gerado.
        pai_is_prensado (bool): Se True, todos os blocos herdam comportamento de prensado.

    Efeitos colaterais:
        Cria um arquivo .xlsm em `pasta`. Usa um arquivo temporário (.tmp)
        para evitar corrupção em caso de falha durante a escrita.
    """
    # ===== CARREGAMENTO DO MOLDE =====
    wb = load_workbook(molde, keep_vba=True)  # keep_vba preserva as macros do molde
    ws = wb.active

    # Conta total de linhas necessárias (+1 por bloco prensado = linha de cabeçalho)
    total_itens = sum(len(b['itens']) + (1 if b['tipo'] == 'prensado' else 0) for b in blocos)
    l_obs = ajustar_molde_elastico(ws, total_itens)

    # ===== DEFINIÇÃO DE ESTILOS =====
    fill_botao = PatternFill(start_color='0078D7', end_color='0078D7', fill_type='solid')  # Azul — botões +5 e ⇄
    fill_erro  = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Rosa — quantidade zero
    font_botao   = Font(name='Arial', color='FFFFFF', bold=True, size=10)  # Texto dos botões azuis
    font_veio    = Font(name='Arial', color='FFFFFF', bold=True, size=8)   # Símbolo de veio ⇄
    font_arial_14      = Font(name='Arial', size=12)       # Fonte padrão para A3 (quantidade)
    font_arial_14_bold = Font(name='Arial', size=12, bold=True)  # Fonte para B3 (título)
    align_botao  = Alignment(horizontal='center', vertical='center')
    align_centro = Alignment(horizontal='center', vertical='center')  # Alinhamento central para material e processo
    font_material = Font(name='Arial', size=8)  # Fonte reduzida para col 9 (material)
    font_processo = Font(name='Arial', size=9)  # Fonte reduzida para col 11 (processo)

    # ===== CABEÇALHO DA PLANILHA =====
    cod_p, acab_p, desc_p = limpar(pai[1]), limpar(pai[2]), limpar(pai[3])

    # Título: inclui acabamento se existir (ex: "1001_BCO - BALCÃO"), senão sem ele
    tit_base = (
        f"{cod_p}_{acab_p.strip(' _-')} - {desc_p}"
        if acab_p.strip(' _-')
        else f"{cod_p} - {desc_p}"
    )
    tit = tit_base

    tratar_cabecalho_a1(ws, id_proj)  # Logo ou nome da marca em A1

    escrever_seguro(ws, 'B3', tit)    # Título do plano de corte em B3
    ws['B3'].font = font_arial_14_bold

    # Quantidade total em A3 — usada como multiplicador nas fórmulas da coluna A
    try:
        ws['A3'].value = float(str(qtd_tot).replace(',', '.'))
    except:
        ws['A3'].value = qtd_tot
    ws['A3'].font = font_arial_14

    # Data de geração em M2
    ws['M2'].value = datetime.now().strftime('%d/%m/%Y')
    ws['M2'].font = Font(name='Arial', size=12)

    # Listas de materiais especiais carregadas das regras
    mat_veio = REGRAS["especiais"]["materiais_com_veio"]    # Materiais que ativam o botão de veio ⇄
    mat_esp  = REGRAS["especiais"]["materiais_plus_5mm"]    # Materiais que recebem +5mm (KRION, DURASEIN, etc.)

    # ===== ITERAÇÃO SOBRE OS BLOCOS DE ITENS =====
    row_idx = 6  # Primeira linha de dados no molde

    for b in blocos:
        # Um bloco herda status de prensado se o pai for prensado OU se o próprio bloco for
        bloco_e_prensado = (b['tipo'] == 'prensado' or pai_is_prensado)

        # --- ORDENAÇÃO POR ESPESSURA (Apenas itens normais, Crescente) ---
        # Ordena para que peças mais finas apareçam primeiro na planilha
        if b['tipo'] == 'normal':
            b['itens'].sort(key=lambda i: buscar_valor_valido(i, [12, 13]))
        # -----------------------------------------------------------------

        # ===== LINHA DE CABEÇALHO DO BLOCO PRENSADO =====
        if b['tipo'] == 'prensado':
            ws.row_dimensions[row_idx].height = 25.5
            # Mescla colunas B até M para o cabeçalho do prensado
            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=13)
            cell_h = ws.cell(row=row_idx, column=2)
            cell_h.value = f"{limpar(b['prensado_info'][1])} - {limpar(b['prensado_info'][3])}"
            cell_h.font = Font(name='Arial', bold=True, size=10)
            cell_h.alignment = align_botao
            row_idx += 1

        # ===== ESCRITA DE CADA ITEM =====
        for item in b['itens']:
            r = row_idx
            is_mig = item.get('is_migrado', False)  # True se veio de migração de planilha antiga
            desc_f = str(limpar(item.get(3, "")))

            # --- EXTRAÇÃO DE DADOS: MIGRADO vs NORMAL ---
            if is_mig:
                # Item migrado: usa dados extraídos diretamente do arquivo antigo
                d_l  = item['desc_orig']   # Descrição original da peça
                m_l  = item['mat_orig']    # Material original
                fita = str(item['fita_orig'])
                veio = converter_para_numero(item.get('veio_orig'))
                fita_b = item.get('fita_lat')  # Fita lateral: '-' ou '='
                fita_e = item.get('fita_top')  # Fita topo: '-' ou '='
                txt = f"{d_l} {m_l}".upper()   # Texto composto para detecção de regras
            else:
                # Item normal: monta o texto composto para aplicar regras
                txt = f"{str(item.get(2, ''))} {desc_f} {str(item.get(14, ''))}".upper()

                # Separa descrição em d_l (parte antes do " - ") e m_l (material)
                d_l, m_l = (desc_f.split(" - ", 1) if " - " in desc_f else ("-", desc_f))

                # Materiais especiais (KRION, DURASEIN...): d_l recebe o código MP (col 14)
                if any(m in txt for m in mat_esp):
                    d_l = str(limpar(item.get(14, "")))

                # Fitas de borda: cols 15 (lateral) e 16 (topo) do PDM
                fita_col_b = str(limpar(item.get(15, "")))
                fita_col_e = str(limpar(item.get(16, "")))
                fita_b = fita_col_b if fita_col_b in ['-', '='] else None
                fita_e = fita_col_e if fita_col_e in ['-', '='] else None

                # Detecção do processo de corte com base no material e fitas
                _madeira_bruta = any(m in txt for m in ["MADEIRA BRUTA PINUS", "MADEIRA BRUTA TAUARI"])
                if _madeira_bruta and (fita_b or fita_e):
                    fita = "SERRA-LAM"
                elif _madeira_bruta:
                    fita = "SERRA"
                elif fita_b or fita_e:
                    fita = "SEC-LAM"
                else:
                    fita = "SEC"
                veio = None

            # --- SUFIXO DIMENSIONAL PARA PRENSADOS ---
            # Bloco prensado: col 12 recebe as dimensões consolidadas (CxLxE)
            if bloco_e_prensado and not is_mig:
                _suf = _sufixo_dimensional(b['itens'], b.get('prensado_info', pai))
                d_l_final = _suf.strip() if _suf else d_l
            else:
                d_l_final = d_l

            # --- REGRA DE INCREMENTO DIMENSIONAL ---
            # FORMICA: +25mm em comp e larg (margem maior para fixação)
            # Materiais especiais (KRION, DURASEIN...): +5mm padrão
            # Migrados e prensados: sem incremento
            _is_formica = not is_mig and not bloco_e_prensado and ('FORM' in txt or 'FORMICA' in txt)
            if _is_formica:
                plus = 25
            elif any(m in txt for m in mat_esp) and not is_mig and not bloco_e_prensado:
                plus = 5
            else:
                plus = 0

            val_fat = float(item.get('q_unitaria_fatorada', 0))

            # ===== ESCRITA DAS CÉLULAS =====

            # Col A: fórmula Excel que multiplica o fator pelo A3 (quantidade do pai)
            ws.cell(row=r, column=1).value = f"={val_fat}*A3"

            # --- FITAS DE BORDA (cols B e E) ---
            font_fita = Font(name="Arial", size=12, bold=True)
            if fita_b:
                c = ws.cell(row=r, column=2)
                c.value, c.font = fita_b, font_fita
                c.data_type = 's'  # Força string — evita que "=" seja interpretado como fórmula
            if fita_e:
                c = ws.cell(row=r, column=5)
                c.value, c.font = fita_e, font_fita
                c.data_type = 's'

            # --- DIMENSÕES: busca com fallback entre colunas ---
            # Comprimento: prioridade col 15 (FB) → col 8 → col 9
            v_c = buscar_valor_valido(item, [15, 8, 9])
            # Largura: prioridade col 16 (FB) → col 10 → col 11
            v_l = buscar_valor_valido(item, [16, 10, 11])
            # Espessura: prioridade col 12 → col 13
            v_a = buscar_valor_valido(item, [12, 13])

            # Aplica o incremento dimensional após encontrar os valores
            if v_c > 0: v_c += plus
            if v_l > 0: v_l += plus

            ws.cell(row=r, column=3).value = v_c  # Comprimento
            ws.cell(row=r, column=6).value = v_l  # Largura
            ws.cell(row=r, column=8).value = v_a  # Espessura

            # --- REGRAS DE DESCRIÇÃO (col 12) ---
            # FORMICA ou VIROLA FLEX: o código MP (col 14) substitui a descrição
            # Isso permite identificar a chapa específica no plano de corte
            if not is_mig and ('FORM' in txt or 'FORMICA' in txt or m_l.upper().startswith('COMPENSADO VIROLA FLEX')):
                _mp_desc = str(limpar(item.get(14, '')))
                if _mp_desc:
                    d_l_final = _mp_desc
            ws.cell(row=r, column=12).value = d_l_final

            # Col 9: material limpo (sem sujeiras do PDM), fonte reduzida e centralizado
            ws.cell(row=r, column=9).value = limpar_material_rigoroso(m_l)
            ws.cell(row=r, column=9).font = font_material
            ws.cell(row=r, column=9).alignment = align_centro

            # --- VEIO (col 10) ---
            if is_mig:
                # Migrado: propaga o valor original (1 = tem veio, None = sem veio)
                ws.cell(row=r, column=10).value = 1 if veio == 1 else None
            else:
                # Normal: verifica se o material está na lista de materiais com veio
                tem_v = any(m in txt for m in mat_veio) or (
                    "KRION" in txt and str(converter_para_numero(item.get(12, ""))) == "3"
                )
                if tem_v:
                    ws.cell(row=r, column=10).value = 1

            # Col 11: processo de corte (SEC, SEC-LAM, SERRA...), fonte e alinhamento
            ws.cell(row=r, column=11).value, ws.cell(row=r, column=11).font, \
                ws.cell(row=r, column=11).alignment = fita, font_processo, align_centro

            # Col 13: código do item — pode ser numérico (1147128) ou alfanumérico (PÇ 1)
            # remover_unidades=False evita que "PÇ 1" seja interpretado incorretamente
            _cod_raw = item.get(1, "")
            _cod_num = converter_para_numero(_cod_raw, remover_unidades=False)
            ws.cell(row=r, column=13).value = _cod_num if _cod_num is not None else limpar(_cod_raw) or None

            # --- BOTÃO DE VEIO ⇄ (col 15) ---
            # Aparece quando a célula de veio (col 10) for 1
            if ws.cell(row=r, column=10).value == 1:
                cv = ws.cell(row=r, column=15)
                cv.value, cv.fill, cv.font, cv.alignment = "⇄", fill_botao, font_veio, align_botao

            # --- BOTÃO +5 (col 14) ---
            # Exibido apenas para itens sem material especial e fora de blocos prensados
            # FORMICA não recebe o botão +5 (já tem +25 implícito)
            if not bloco_e_prensado and not any(m in txt for m in mat_esp):
                cn = ws.cell(row=r, column=14)
                cn.value, cn.fill, cn.font, cn.alignment = "+5", fill_botao, font_botao, align_botao

            row_idx += 1

    # ===== RODAPÉ: PROJETO DE REFERÊNCIA =====
    escrever_seguro(ws, f"A{l_obs}", f"PROJETO DE REFERÊNCIA: {id_proj}", Alignment(horizontal='left'))
    ws[f"A{l_obs}"].font = Font(name='Arial Black', size=14, bold=True)

    # ===== SALVAMENTO SEGURO VIA ARQUIVO TEMPORÁRIO =====
    # Remove caracteres inválidos em nomes de arquivo do Windows
    nome_base_limpo = re.sub(r'[\\/*?:\u0022<>|]', '', tit).strip()[:120]
    caminho = os.path.join(pasta, f"{nome_base_limpo}.xlsm")

    # Salva em .tmp primeiro; só substitui o definitivo após sucesso total
    caminho_tmp = caminho + ".tmp"
    try:
        wb.save(caminho_tmp)
        if os.path.exists(caminho):
            os.remove(caminho)         # Remove versão anterior se existir
        os.rename(caminho_tmp, caminho)  # Promove .tmp para o arquivo final
    except Exception as e:
        if os.path.exists(caminho_tmp):
            os.remove(caminho_tmp)     # Limpa arquivo temporário em caso de falha
        raise e