"""
================================================================================
core/migracao.py — Migração de Planilhas Antigas (ODS / XLSX)
================================================================================
Responsável por:
  1. Extrair dados de planilhas antigas (.ods e .xlsx) para o formato interno
  2. Verificar duplicidade de códigos na rede (cache de arquivos)
  3. Mapear códigos MP para nomes de material atualizados (mapeamento.json)

Sistema de Extração Inteligente:
  - Detecta automaticamente o layout da planilha (com ou sem coluna UN)
  - Usa um sistema de "fila" para resistir a colunas apagadas ou deslocadas
  - Suporta blocos prensados dentro de planilhas antigas

Fluxo:
  extrair_dados_migracao()
    → lê .ods ou .xlsx
    → para cada linha: extrair_dados_linha_inteligente()
    → retorna lista de blocos com itens no formato padrão do sistema
================================================================================
"""

# ===== IMPORTAÇÕES =====
import os
import re
import json
from pathlib import Path

import pandas as pd              # Leitura de .ods (engine='odf')
from openpyxl import load_workbook  # Leitura de .xlsx/.xlsm

from utils.config import REGRAS  # Dicionário global de regras (regras.json)
from utils.helpers import limpar  # Normaliza strings vindas do Excel


# ===== CARREGAMENTO DO MAPEAMENTO MP → MATERIAL =====

_MAP_MP_PATH = Path(__file__).parent.parent / "mapeamento.json"

try:
    with open(_MAP_MP_PATH, encoding='utf-8') as _f:
        _MAP_MP = json.load(_f)
    # Ex: {"915239": "COMPENSADO VIROLA FLEX BALCAO", "925324": "KRION ..."}
except Exception:
    _MAP_MP = {}  # Fallback silencioso — o sistema continua sem mapeamento


# ===== FUNÇÕES AUXILIARES DE DETECÇÃO =====

def is_pdm_code(s):
    """
    Verifica se uma string parece ser um código válido do PDM.

    Utilizado para distinguir o código da peça (col M) de textos descritivos
    que podem ter migrado para a mesma posição em planilhas com colunas deslocadas.

    Parâmetros:
        s (str): String a avaliar.

    Retorna:
        bool: True se parece ser um código PDM válido.

    Exemplos aceitos:
        "1147128" → True (código numérico Ingecon)
        "PÇ 1"   → True (código alfanumérico de peça)
        "MDF CRU" → False (nome de material)
    """
    s = str(s).strip().upper()
    if not s:
        return False

    # Aceita variações de "PÇ 1", "PC 12", "PÇ" — códigos alfanuméricos de peça
    if re.match(r'^P[CÇ]\s*\d*$', s):
        return True

    # Aceita códigos Ingecon: 11xxxxx, 15xxxxx, 0xxxxx (com sufixo de letra opcional)
    if re.match(r'^(11|15|0)\d{4,}[A-Z]?$', s):
        return True

    # Aceita códigos MP numéricos muito grandes (geralmente > 50.000)
    try:
        val = float(s.replace(',', '.'))
        if val > 50000:
            return True
    except:
        pass

    return False


# ===== EXTRAÇÃO INTELIGENTE DE LINHA =====

def extrair_dados_linha_inteligente(linha_raw, a3_valor, cod_fallback=''):
    """
    Extrai os dados de uma linha de planilha antiga de forma tolerante a variações de layout.

    O sistema usa âncoras (comprimento >= 50mm) para determinar o offset de colunas,
    permitindo que planilhas com colunas apagadas ou deslocadas sejam lidas corretamente.

    Parâmetros:
        linha_raw (list): Valores brutos da linha da planilha.
        a3_valor (float): Quantidade total do projeto (divisor para calcular fator unitário).
        cod_fallback (str): Código a usar se nenhum código for encontrado na linha.

    Retorna:
        dict | None: Dicionário com os campos do item (comp, larg, esp, mat, fita...),
                     ou None se a linha não contiver dados válidos.
    """
    linha = [limpar(x) for x in linha_raw]

    # Garante que a linha tem colunas suficientes para o mapeamento
    while len(linha) < 25:
        linha.append("")

    # --- PASSO 1: ENCONTRAR A ÂNCORA (Comprimento) ---
    # O comprimento é a primeira dimensão >= 50mm, geralmente na col 2 ou 3
    # O offset permite compensar planilhas onde a col A foi deletada
    comp_idx = -1
    for i in range(5):
        val = re.sub(r'(?i)\s*(pç|pc|un|und|unid|unidade)\.*', '', str(linha[i])).strip()
        try:
            if float(val.replace(',', '.')) >= 50:
                comp_idx = i
                break
        except:
            pass

    if comp_idx == -1:
        return None  # Linha sem dimensão válida — provavelmente uma linha de cabeçalho

    # Offset: diferença entre posição encontrada e posição esperada (col 2 = índice 2)
    offset = comp_idx - 2

    # --- PASSO 2: EXTRAÇÃO DE QUANTIDADE ---
    qtd_idx = 0 + offset
    if qtd_idx >= 0:
        qtd_str = re.sub(r'(?i)\s*(pç|pc|un|und|unid|unidade)\.*', '', str(linha[qtd_idx])).strip()
        try:
            q_val = float(qtd_str.replace(',', '.'))
        except:
            q_val = 0
    else:
        q_val = 0

    # Fator unitário: divide pela quantidade do projeto para obter proporção por peça
    q_unit = q_val / a3_valor if a3_valor > 0 else q_val

    # Verifica se a col 1 tem um fator direto (< 50, positivo) — planilhas com coluna UN
    if 1 + offset >= 0:
        q1_str = re.sub(r'(?i)\s*(pç|pc|un|und|unid|unidade)\.*', '', str(linha[1 + offset])).strip()
        try:
            q1_val = float(q1_str.replace(',', '.'))
            if q1_val > 0 and q1_val < 50:
                q_unit = q1_val  # Col 1 é um fator unitário direto (não dividir por a3)
        except:
            pass

    # --- EXTRAÇÃO DAS DIMENSÕES E FITAS DE BORDA ---
    comp     = linha[2 + offset]  # Comprimento
    larg     = linha[5 + offset]  # Largura
    esp      = linha[7 + offset]  # Espessura

    # Fitas de borda: '-' ou '=' nas colunas vizinhas às dimensões
    fita_lat = linha[1 + offset] if linha[1 + offset] in ['-', '='] else None  # Lateral
    fita_top = linha[4 + offset] if linha[4 + offset] in ['-', '='] else None  # Topo

    # --- PASSO 3: EXTRAÇÃO GEOGRÁFICA (Colunas I, J, K, L, M) ---
    mat      = str(linha[8 + offset]).strip()   # Col I: material
    veio_val = str(linha[9 + offset]).strip().upper()  # Col J: veio ou processo

    # Termos que indicam que a coluna contém o PROCESSO (não o veio)
    termos_proc = ['SEC', 'SEC-LAM', 'SERRA', 'SERRA-LAM', 'LAM', 'FITA', 'BORDA', 'BORD', 'USI', 'USINAGEM', 'PRENSA', 'CNC', 'MAR']
    is_proc = lambda x: any(t in x.upper() for t in termos_proc)

    fita = ""
    desc = ""
    cod  = ""

    # --- SISTEMA DE FILA: resiste a colunas apagadas ---
    if is_proc(veio_val):
        # Col J foi identificada como Processo — Veio foi deletado ou inexistente
        veio = None
        fita = veio_val
        # Fila começa em K (índice 10) pois Processo consumiu a posição do Veio
        fila = [str(linha[i + offset]).strip() for i in range(10, 14)]
    else:
        # Col J é o Veio: 1 = tem veio, None = sem veio
        veio = 1 if veio_val in ['1', '1.0', 'S', 'SIM', 'X'] else None
        fila = [str(linha[i + offset]).strip() for i in range(10, 15)]

        if is_proc(fila[0]):
            fita = fila.pop(0)   # Primeiro elemento da fila é o Processo
        elif fila[0] == "":
            fita = fila.pop(0)   # Processo estava vazio — consome o slot mesmo assim

    # O que sobra na fila segue a ordem: Descrição (col L) → Código (col M)
    desc = fila.pop(0)
    c1   = fila.pop(0)
    c2   = fila.pop(0)

    # --- PASSO 4: IDENTIFICAÇÃO DE CÓDIGO vs DESCRIÇÃO ---
    # c2 tem prioridade se parece código PDM; senão tenta c1
    if is_pdm_code(c2):
        cod  = c2
        desc = f"{desc} {c1}".strip()  # c1 é extensão da descrição
    elif is_pdm_code(c1):
        cod  = c1
        if c2:
            desc = f"{desc} {c2}".strip()
    else:
        cod  = c1
        if c2:
            desc = f"{desc} {c2}".strip()

    # --- PASSO 5: FALLBACK PARA CÓDIGO EMBUTIDO NA DESCRIÇÃO ---
    # Planilhas com colunas mescladas podem ter "DESCRIÇÃO - 1147128" na mesma célula
    if " - " in desc and not is_pdm_code(cod):
        parts = desc.rsplit(" - ", 1)
        if is_pdm_code(parts[1]):
            desc, cod = parts[0].strip(), parts[1].strip()
        elif is_pdm_code(parts[0]):
            desc, cod = parts[1].strip(), parts[0].strip()

    # Prevenção: se desc parece ser um código e cod está vazio, inverte
    if is_pdm_code(desc) and not is_pdm_code(cod):
        cod  = desc
        desc = ""

    # Usa o código do título da planilha como fallback se nenhum foi encontrado
    if not cod:
        cod = cod_fallback

    return {
        1:  cod,           # Código do item
        8:  comp,          # Comprimento
        10: larg,          # Largura
        12: esp,           # Espessura
        'mat_orig':           mat,       # Material original
        'veio_orig':          veio,      # 1 = tem veio, None = sem veio
        'fita_orig':          fita,      # Processo de corte (SEC, SEC-LAM...)
        'fita_lat':           fita_lat,  # Fita de borda lateral ('-' ou '=')
        'fita_top':           fita_top,  # Fita de borda topo ('-' ou '=')
        'desc_orig':          desc,      # Descrição da peça
        'q_unitaria_fatorada': q_unit,   # Fator unitário (quantidade por peça)
        'is_migrado':         True       # Marca o item como migrado
    }


# ===== EXTRAÇÃO PRINCIPAL =====

def extrair_dados_migracao(caminho):
    """
    Lê uma planilha antiga (.ods ou .xlsx) e extrai os dados no formato interno do sistema.

    Suporta dois formatos de arquivo:
      - ODS: planilhas antigas abertas no LibreOffice (leitura via pandas+odf)
      - XLSX/XLSM: planilhas antigas no formato Excel (leitura via openpyxl)

    Parâmetros:
        caminho (str | Path): Caminho completo para o arquivo a migrar.

    Retorna:
        tuple(list[dict], float):
          - Lista de blocos {'tipo': 'normal'|'prensado', 'itens': [...]}
          - Valor de a3 (quantidade total do projeto no arquivo antigo)
          Em caso de erro: ([], 1.0)
    """
    try:
        # Termos que identificam linhas de cabeçalho a ignorar
        termos_ignorar = ["PROGRAMAÇÃO", "DATA", "UN", "MEDIDA", "MATERIAL", "PROCESSO", "DESCRIÇÃO", "CÓDIGO", "QNT", "PROG."]

        blocos      = []
        bloco_atual = {'tipo': 'normal', 'itens': []}
        linhas_raw  = []
        a3_valor    = 1.0   # Quantidade padrão se não encontrar na planilha
        cod_titulo  = ''    # Código extraído do título (fallback para itens sem código)

        # ===== LEITURA DO ARQUIVO =====
        if str(caminho).lower().endswith('.ods'):
            # --- Leitura ODS via pandas ---
            df_old = pd.read_excel(caminho, engine='odf', header=None).fillna('')
            while df_old.shape[1] < 16:
                df_old[df_old.shape[1]] = ''  # Garante colunas suficientes

            # a3: primeira célula numérica positiva na linha 3 (índice 2)
            linha_a3 = [limpar(df_old.iloc[2, c]) for c in range(df_old.shape[1])]
            for cell in linha_a3:
                try:
                    num = float(str(cell).replace(',', '.'))
                    if num > 0:
                        a3_valor = num
                        break
                except:
                    pass

            # Código do título: está na linha 2, na célula com " - " (ex: "1147128 - BALCÃO")
            linha_titulo = [limpar(df_old.iloc[1, c]) for c in range(df_old.shape[1])]
            for cell in linha_titulo:
                if " - " in cell:
                    cod_titulo = cell.split(' - ')[0].strip()
                    break

            # Coleta todas as linhas de dados (a partir da linha 6, índice 5)
            for r in range(5, len(df_old)):
                linhas_raw.append([df_old.iloc[r, c] for c in range(df_old.shape[1])])

        else:
            # --- Leitura XLSX/XLSM via openpyxl ---
            ws_d = load_workbook(caminho, data_only=True).active

            # a3: tenta A3 primeiro, depois varre a linha 3 em busca de número positivo
            try:
                a3_valor = float(str(ws_d['A3'].value).replace(',', '.'))
            except:
                a3_valor = 0

            if a3_valor == 0:
                for c in range(1, 10):
                    try:
                        num = float(str(ws_d.cell(row=3, column=c).value).replace(',', '.'))
                        if num > 0:
                            a3_valor = num
                            break
                    except:
                        pass

            a3_valor = 1.0 if a3_valor == 0 else a3_valor

            # Coleta até 500 linhas (limite para evitar planilhas corrompidas)
            for r in range(1, min(500, ws_d.max_row + 1)):
                linhas_raw.append([ws_d.cell(row=r, column=c).value for c in range(1, 20)])

        # Gatilhos que identificam linhas de cabeçalho de bloco prensado
        gatilhos_pren = REGRAS["prensados"]["descricoes_gatilho"]  # Ex: ["PRENSADO"]
        codigos_pren  = REGRAS["prensados"]["codigos_gatilho"]      # Ex: ["1152032"]

        # ===== PROCESSAMENTO LINHA A LINHA =====
        for linha_bruta in linhas_raw:
            linha = [limpar(x) for x in linha_bruta]
            while len(linha) < 16:
                linha.append('')

            texto_linha = " ".join([str(x) for x in linha if x]).upper()
            if not texto_linha:
                continue  # Pula linhas completamente vazias

            # --- FILTRO DE LINHAS DE CABEÇALHO ---
            # Termos curtos (UN, QNT) checados por igualdade; termos longos por substring
            _termos_longos = [t for t in termos_ignorar if len(t) >= 4]
            ignorar = False
            for cell in linha:
                cell_up = str(cell).upper()
                if cell_up in ["UN", "QNT", "QTD"] or any(t in cell_up for t in _termos_longos):
                    ignorar = True
                    break
            if ignorar:
                continue

            # --- DETECÇÃO DE BLOCO PRENSADO ---
            if any(g in texto_linha for g in gatilhos_pren) or any(c in texto_linha for c in codigos_pren):
                # Fecha o bloco atual antes de abrir o novo
                if bloco_atual['itens']:
                    blocos.append(bloco_atual)

                # Extrai o texto do prensado (célula que contém o gatilho)
                texto_prensado = ""
                for cell in linha:
                    cell_up = str(cell).upper()
                    if any(g in cell_up for g in gatilhos_pren) or any(c in cell_up for c in codigos_pren):
                        texto_prensado = str(cell)
                        break
                if not texto_prensado:
                    for cell in linha:
                        if cell:
                            texto_prensado = str(cell)
                            break

                # Separa código e descrição do prensado (se no formato "COD - DESC")
                f_cod, f_desc = "", texto_prensado
                if " - " in texto_prensado:
                    partes = texto_prensado.split(" - ", 1)
                    f_cod, f_desc = partes[0].strip(), partes[1].strip()

                bloco_atual = {'tipo': 'prensado', 'prensado_info': {1: f_cod, 3: f_desc}, 'itens': []}
                continue

            # --- EXTRAÇÃO DO ITEM ---
            item = extrair_dados_linha_inteligente(linha, a3_valor, cod_titulo)
            if item:
                bloco_atual['itens'].append(item)

        # Fecha o último bloco
        if bloco_atual['itens']:
            blocos.append(bloco_atual)

        return blocos, a3_valor

    except Exception as e:
        print(f"[MIGRATION ERROR] {e}")
        return [], 1.0  # Retorno seguro — processador tentará processar do zero


# ===== CACHE E VERIFICAÇÃO DE DUPLICIDADE NA REDE =====

def mapear_rede_cache():
    """
    Mapeia todos os arquivos Excel/ODS na pasta raiz da rede em uma lista de tuplas.

    Varre recursivamente a pasta pai de PLANOS DE CORTE 2026, que inclui tanto
    os arquivos novos (2026) quanto os antigos (outras subpastas).

    Usa lista em vez de dicionário para evitar sobreposição quando dois arquivos
    têm o mesmo nome em pastas diferentes.

    Retorna:
        list[tuple(str, str)]: Lista de (nome_do_arquivo, caminho_completo).
    """
    cache = []
    pasta_base = Path(REGRAS["diretorios"]["raiz"]).parent  # Sobe um nível acima de PLANOS DE CORTE 2026

    if pasta_base.exists():
        for root, _, files in os.walk(pasta_base):
            for f in files:
                # Inclui apenas formatos de planilha suportados
                if f.lower().endswith(('.xlsx', '.ods', '.xlsm')):
                    cache.append((f, os.path.join(root, f)))

    return cache


def verificar_duplicidade_em_rede(codigo, cache_rede):
    """
    Verifica se um código de peça já existe como arquivo na rede.

    A busca usa regex com boundary para evitar falsos positivos:
    "1147" não casa com "11470001". O código deve ser seguido de
    separador válido (espaço, hífen, underscore, ponto) ou fim de string.

    Parâmetros:
        codigo (str): Código da peça a buscar (ex: "1147128").
        cache_rede (list): Cache retornado por mapear_rede_cache().

    Retorna:
        str | None: Caminho completo do arquivo encontrado, ou None se não existir.
    """
    c = str(codigo).strip()
    if not c:
        return None

    # Regex: código seguido de separador válido ou fim de nome de arquivo
    padrao = re.compile(rf"^{re.escape(c)}(\s|-|_|\.|$)")

    for f, caminho in cache_rede:
        if padrao.match(f):
            return caminho  # Retorna o primeiro match encontrado

    return None