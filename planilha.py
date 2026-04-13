# =============================================================================
# BIBLIOTECAS E DEPENDÊNCIAS
# =============================================================================
import os
import sys
import shutil
import re
import threading
from pathlib import Path
from datetime import datetime
from copy import copy

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment, PatternFill

import tkinter as tk
import customtkinter as ctk
from tkinter import messagebox

# =============================================================================
# CONFIGURAÇÕES GLOBAIS E DIRETÓRIOS
# =============================================================================
VERSAO_ATUAL = "2.2.04"

# Caminhos de Rede
DIRETORIO_RAIZ_PLANILHAS = Path(r"X:\Egpe\06 - PLANOS DE CORTE ATUALIZADOS\PLANOS DE CORTE 2026")
DIRETORIO_ANTIGOS = Path(r"X:\Egpe\06 - PLANOS DE CORTE ATUALIZADOS\ANTIGOS - NÃO USAR")
DIRETORIO_SISTEMA = DIRETORIO_RAIZ_PLANILHAS / "GeradorPlanilhasAutomação"

PASTAS_VERIFICACAO = [
    Path(r"X:\Egpe\06 - PLANOS DE CORTE ATUALIZADOS")
]

# Arquivos de Suporte
ARQ_VERSAO = DIRETORIO_SISTEMA / "version.txt"
EXE_SERVIDOR = DIRETORIO_SISTEMA / "Gerador_Planilhas_Ingecon.exe"

# Caminhos Locais
DESKTOP_PATH = Path(os.path.join(os.path.expanduser("~"), "Desktop"))
ARQ_REPETIDAS = DESKTOP_PATH / "planilhas_repetidas.txt"

# Identidade Visual (Cores)
COR_PRINCIPAL = "#d32732"
COR_HOVER = "#a81f28"
COR_TESTE = "#e67e22" 

# Configuração de Aparência do CustomTkinter
ctk.set_appearance_mode("Light") 

# =============================================================================
# CLASSE PRINCIPAL DA APLICAÇÃO
# =============================================================================
class AppIngecon(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        # Inicialização do Sistema
        self.verificar_atualizacao()
        self.setup_window()
        self.setup_variables()
        self.create_widgets()
        
        # Atalhos de Teclado
        self.bind("<Key>", self.verificar_codigo_secreto)

    # -------------------------------------------------------------------------
    # CONFIGURAÇÃO DE INTERFACE (UI)
    # -------------------------------------------------------------------------
    def setup_window(self):
        """Configurações básicas da janela principal."""
        self.title(f"Ingecon - Gerador de Planilhas V{VERSAO_ATUAL}")
        self.geometry("480x360") 
        self.configure(fg_color="#f5f5f5") 
        self.grid_columnconfigure(0, weight=1)

    def setup_variables(self):
        """Variáveis de controle de estado."""
        self.modo_teste_ativo = False
        self.buffer_teclas = ""
        self.SECRET_CODE = "dev"

    def create_widgets(self):
        """Criação e posicionamento dos elementos visuais."""
        # Header (Cabeçalho)
        self.header_frame = ctk.CTkFrame(self, fg_color=COR_PRINCIPAL, height=70, corner_radius=0)
        self.header_frame.grid(row=0, column=0, sticky="nsew", pady=(0, 20))
        self.header_frame.grid_columnconfigure(0, weight=1)

        self.label_titulo = ctk.CTkLabel(
            self.header_frame, 
            text="GERADOR DE PLANILHAS", 
            font=ctk.CTkFont(size=22, weight="bold"), 
            text_color="white"
        )
        self.label_titulo.grid(row=0, column=0, pady=20)

        # Botão Principal de Processamento
        self.btn_processar = ctk.CTkButton(
            self, 
            text="Colar - Gerar Planilhas", 
            command=self.iniciar_processamento,
            fg_color=COR_PRINCIPAL, 
            hover_color=COR_HOVER, 
            height=50, 
            corner_radius=8, 
            font=ctk.CTkFont(size=15, weight="bold")
        )
        self.btn_processar.grid(row=1, column=0, padx=40, pady=40)

        # Barra de Progresso (Inicia oculta)
        self.progress = ctk.CTkProgressBar(
            self, 
            orientation="horizontal", 
            progress_color=COR_PRINCIPAL, 
            width=300
        )
        self.progress.set(0)

    # -------------------------------------------------------------------------
    # ÁREA DE SEGURANÇA E MODO DESENVOLVEDOR
    # -------------------------------------------------------------------------
    def verificar_codigo_secreto(self, event):
        """Verifica sequência de teclas para ativar o modo de teste."""
        self.buffer_teclas += event.char.lower()
        if len(self.buffer_teclas) > len(self.SECRET_CODE):
            self.buffer_teclas = self.buffer_teclas[-len(self.SECRET_CODE):]
            
        if self.buffer_teclas == self.SECRET_CODE:
            self.modo_teste_ativo = not self.modo_teste_ativo
            self.buffer_teclas = ""
            self.atualizar_visual_teste()

    def atualizar_visual_teste(self):
        """Altera as cores da interface se o modo teste estiver ativo."""
        if self.modo_teste_ativo:
            self.btn_processar.configure(fg_color=COR_TESTE, hover_color="#d35400", text="MODO TESTE ATIVO")
        else:
            self.btn_processar.configure(fg_color=COR_PRINCIPAL, hover_color=COR_HOVER, text="Colar - Gerar Planilhas")

    # -------------------------------------------------------------------------
    # SISTEMA DE ATUALIZAÇÃO AUTOMÁTICA
    # -------------------------------------------------------------------------
    def verificar_atualizacao(self):
        """Checa se existe uma versão mais recente no servidor."""
        try:
            if ARQ_VERSAO.exists():
                with open(ARQ_VERSAO, "r") as f: 
                    v_serv = f.read().strip()
                if v_serv != VERSAO_ATUAL:
                    if messagebox.askyesno("Atualização", f"Nova versão {v_serv} disponível. Atualizar?"):
                        self.executar_patch()
                        os._exit(0)
        except Exception as e: 
            print(f"[UPDATE WARN] {e}")

    def executar_patch(self):
        """Gera e executa um script batch para substituir o executável atual."""
        c_exe, n_exe = sys.executable, os.path.basename(sys.executable)
        bat = (
            f'@echo off\n:loop\ntaskkill /f /im "{n_exe}" >nul 2>&1\n'
            f'del /f /q "{c_exe}" >nul 2>&1\n'
            f'if exist "{c_exe}" (timeout /t 1 >nul\ngoto loop)\n'
            f'copy /y "{EXE_SERVIDOR}" "{c_exe}"\nstart "" "{c_exe}"\nexit'
        )
        p_bat = Path(os.environ["TEMP"]) / f"patch_ingecon_{os.getpid()}.bat"
        with open(p_bat, "w") as f: 
            f.write(bat)
        os.startfile(p_bat)

    # -------------------------------------------------------------------------
    # UTILITÁRIOS DE ARQUIVOS E EXCEL
    # -------------------------------------------------------------------------
    def resource_path(self, relative_path):
        """Gerencia caminhos de arquivos internos do executável (PyInstaller)."""
        if hasattr(sys, '_MEIPASS'):
            p_interno = Path(sys._MEIPASS) / relative_path
            if p_interno.exists(): return str(p_interno)
        return str(Path(os.getcwd()) / relative_path)

    def escrever_seguro(self, ws, coord, valor, alinhamento=None):
        """Escreve em células, tratando corretamente células mescladas."""
        try:
            cell = ws[coord]
            if cell.__class__.__name__ == 'MergedCell':
                for r in ws.merged_cells.ranges:
                    if coord in r:
                        m = ws.cell(row=r.min_row, column=r.min_col)
                        m.value = valor
                        if alinhamento: m.alignment = alinhamento
                        return
            else:
                cell.value = valor
                if alinhamento: cell.alignment = alinhamento
        except Exception as e: 
            print(f"[EXCEL WARN] Falha ao escrever na célula {coord}: {e}")

    def limpar(self, val):
        """Limpa strings de dados vindos do Excel/Clipboard."""
        if val is None: return ""
        v = str(val).strip()
        if v.endswith('.0'): v = v[:-2]
        return v if v.lower() not in ['nan', 'none', 'null', ''] else ""

    def converter_para_numero(self, valor):
        """Converte strings para inteiros arredondados."""
        limpo = self.limpar(valor)
        if not limpo or limpo in ["-", "="]: return limpo
        try:
            v_aj = limpo.replace(',', '.')
            val_float = float(v_aj)
            return int(val_float + 0.5) if val_float >= 0 else int(val_float - 0.5)
        except Exception: return limpo

    # -------------------------------------------------------------------------
    # PROCESSAMENTO TÉCNICO DE PLANILHAS
    # -------------------------------------------------------------------------
    def tratar_cabecalho_a1(self, ws, id_projeto):
        """Insere logotipos ou nomes de marcas na célula A1."""
        logos_imagem = {
            "ARO": "amaro", "BGK": "BurguerKing", "CAM": "Camicado", "CEA": "CeA", 
            "CEN": "Centauro", "ELU": "Elubel", "FAR": "FarmaCopr", "IND": "Indian", 
            "ING": "ingecon", "MCD": "mcdonalds", "PER": "pernambucanas", 
            "REN": "renner", "TMS": "Tramontina", "TRA": "tramontinaPDV"
        }
        marcas_texto = {"ZAR": "ZARA", "ZFR": "ZAFFARI", "SEP": "SEPHORA", "PRO": "PROTÓTIPO"}
        id_up = str(id_projeto).upper()
        ws['A1'].value = None

        for sigla, nome in marcas_texto.items():
            if sigla in id_up:
                self.escrever_seguro(ws, 'A1', nome, Alignment(horizontal='center', vertical='center'))
                ws['A1'].font = Font(size=22, bold=True)
                return

        for sigla, arq in logos_imagem.items():
            if sigla in id_up:
                path = self.resource_path(f"logos/{arq}.png")
                if Path(path).exists():
                    img = OpenpyxlImage(path)
                    img.width, img.height = 152, 42
                    ws.row_dimensions[1].height = 33
                    ws.add_image(img, 'A1')
                    return

        # Padrão Ingecon se não encontrar marca
        path_ing = self.resource_path("logos/ingecon.png")
        if Path(path_ing).exists():
            img = OpenpyxlImage(path_ing)
            img.width, img.height = 152, 42
            ws.row_dimensions[1].height = 33
            ws.add_image(img, 'A1')

    def limpar_material_rigoroso(self, texto):
        """Remove termos técnicos e dimensões da descrição do material."""
        if not texto: return ""
        t = re.sub(r'\b(ORIG|ESS)\b', '', str(texto), flags=re.IGNORECASE).replace('=', '')
        t = re.sub(r'\s*\b\d+(?:[\.,]\d+)?\s*[xX].*$', '', t, flags=re.IGNORECASE)
        return re.sub(r'\s+', ' ', t).strip(' -')

    def ajustar_molde_elastico(self, ws, num_itens):
        """Redimensiona a planilha conforme a quantidade de itens."""
        for r in range(1, 50):
            ws.row_dimensions[r].hidden = False

        padrao, l_rodape, quadro = 3, 9, None
        
        for m in list(ws.merged_cells.ranges):
            if m.min_row >= l_rodape:
                if m.min_row == l_rodape: 
                    quadro = {'min_col': m.min_col, 'max_col': m.max_col, 'max_row': m.max_row}
                try: ws.unmerge_cells(str(m))
                except Exception: pass
                
        if quadro and quadro['max_row'] > l_rodape: 
            ws.delete_rows(l_rodape + 1, quadro['max_row'] - l_rodape)
            
        diff = num_itens - padrao
        if diff > 0: ws.insert_rows(l_rodape, diff)
        elif diff < 0: ws.delete_rows(l_rodape + diff, abs(diff))
        
        for r in range(6, 6 + num_itens):
            ws.row_dimensions[r].height = 25.5
            ws.cell(row=r, column=4).value = ws.cell(row=r, column=7).value = "X"
            if r > 6:
                for c in range(1, 16):
                    src, tgt = ws.cell(row=6, column=c), ws.cell(row=r, column=c)
                    if src.has_style: tgt._style = copy(src._style)
                    
        n_inicio = 6 + num_itens
        if quadro: 
            ws.merge_cells(start_row=n_inicio, start_column=quadro['min_col'], 
                           end_row=n_inicio, end_column=quadro['max_col'])
        return n_inicio

    # -------------------------------------------------------------------------
    # LÓGICA DE MIGRAÇÃO E CACHE DE REDE
    # -------------------------------------------------------------------------
    def mapear_rede_cache(self):
        """Cria um dicionário de arquivos existentes para busca rápida."""
        cache = {}
        for p in PASTAS_VERIFICACAO:
            if not p.exists(): continue
            for root, _, files in os.walk(p):
                for f in files:
                    if f.lower().endswith(('.xlsx', '.ods', '.xlsm')): 
                        cache[f] = os.path.join(root, f)
        return cache

    def verificar_duplicidade_em_rede(self, codigo, cache_rede):
        """Verifica se o código já possui uma planilha gerada."""
        c = str(codigo).strip()
        if not c: return None
        padrao = re.compile(rf"^{re.escape(c)}(\D|$)")
        for f, caminho in cache_rede.items():
            if padrao.match(f): return caminho
        return None

    def extrair_dados_migracao(self, caminho):
        """Extrai dados de planilhas antigas (ODS ou XLSX) para o novo formato."""
        try:
            if str(caminho).lower().endswith('.ods'):
                df_old = pd.read_excel(caminho, engine='odf', header=None).fillna('')
                while df_old.shape[1] < 15: df_old[df_old.shape[1]] = ''
                try: a3_valor = float(self.converter_para_numero(df_old.iloc[2, 0]) or 1.0)
                except Exception: a3_valor = 1.0
                if a3_valor == 0: a3_valor = 1.0
                
                itens = []
                for r in range(5, len(df_old)):
                    cod = self.limpar(df_old.iloc[r, 12])
                    desc = self.limpar(df_old.iloc[r, 11])
                    comp = self.limpar(df_old.iloc[r, 2]); larg = self.limpar(df_old.iloc[r, 5])
                    if not cod and not desc and not comp and not larg: continue
                    if str(cod).upper() == "X" and not desc and not comp and not larg: continue
                    
                    try: fator_bruto = float(self.converter_para_numero(df_old.iloc[r, 0]) or 0)
                    except Exception: fator_bruto = 0.0
                    fator_unitario = fator_bruto / a3_valor if a3_valor > 0 else fator_bruto
                    
                    item = {
                        1: cod, 15: df_old.iloc[r, 1], 8: df_old.iloc[r, 2],
                        16: df_old.iloc[r, 4], 10: df_old.iloc[r, 5],
                        12: df_old.iloc[r, 7], 'mat_orig': df_old.iloc[r, 8],
                        'veio_orig': df_old.iloc[r, 9], 'fita_orig': df_old.iloc[r, 10],
                        'desc_orig': desc, 'q_unitaria_fatorada': fator_unitario, 'is_migrado': True
                    }
                    itens.append(item)
                return itens, a3_valor
            else:
                wb_data = load_workbook(caminho, data_only=True); ws_d = wb_data.active
                try: a3_valor = float(self.converter_para_numero(ws_d['A3'].value) or 1.0)
                except Exception: a3_valor = 1.0
                if a3_valor == 0: a3_valor = 1.0
                
                itens = []
                for r in range(6, 500):
                    cod = self.limpar(ws_d.cell(row=r, column=13).value)
                    desc = self.limpar(ws_d.cell(row=r, column=12).value)
                    comp = self.limpar(ws_d.cell(row=r, column=3).value)
                    larg = self.limpar(ws_d.cell(row=r, column=6).value)
                    if not cod and not desc and not comp and not larg: continue
                    if str(cod).upper() == "X" and not desc and not comp and not larg: continue
                    
                    try: fator_bruto = float(str(ws_d.cell(row=r, column=1).value or 0).replace(',', '.'))
                    except Exception: fator_bruto = 0.0
                    fator_unitario = fator_bruto / a3_valor if a3_valor > 0 else fator_bruto
                    
                    item = {
                        1: cod, 15: ws_d.cell(row=r, column=2).value, 8: ws_d.cell(row=r, column=3).value,
                        16: ws_d.cell(row=r, column=5).value, 10: ws_d.cell(row=r, column=6).value,
                        12: ws_d.cell(row=r, column=8).value, 'mat_orig': ws_d.cell(row=r, column=9).value,
                        'veio_orig': ws_d.cell(row=r, column=10).value, 'fita_orig': ws_d.cell(row=r, column=11).value,
                        'desc_orig': desc, 'q_unitaria_fatorada': fator_unitario, 'is_migrado': True
                    }
                    itens.append(item)
                return itens, a3_valor
        except Exception as e: 
            print(f"[MIGRATION WARN] Erro ao extrair de {caminho}: {e}")
            return [], 1.0

    # -------------------------------------------------------------------------
    # GERAÇÃO DO ARQUIVO FINAL
    # -------------------------------------------------------------------------
    def gerar_arquivo_excel(self, pai, blocos, id_proj, qtd_tot, molde, pasta, pai_is_prensado):
        """Cria o arquivo Excel formatado com os dados coletados."""
        wb = load_workbook(molde, keep_vba=True); ws = wb.active
        total_itens = sum(len(b['itens']) + (1 if b['tipo'] == 'prensado' else 0) for b in blocos)
        
        l_obs = self.ajustar_molde_elastico(ws, total_itens)
        
        # Estilos dos Botões visuais
        fill_botao = PatternFill(start_color='0078D7', end_color='0078D7', fill_type='solid')
        fill_erro = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        font_botao = Font(color='FFFFFF', bold=True, size=10)
        font_veio = Font(color='FFFFFF', bold=True, size=8)
        align_botao = Alignment(horizontal='center', vertical='center')

        # Título do Projeto
        cod_p, acab_p, desc_p = self.limpar(pai[1]), self.limpar(pai[2]), self.limpar(pai[3])
        acab_real = acab_p.strip(" _-")
        tit = f"{cod_p}_{acab_real} - {desc_p}" if acab_real else f"{cod_p} - {desc_p}"
        
        self.tratar_cabecalho_a1(ws, id_proj)
        self.escrever_seguro(ws, 'B3', tit)
        try: ws['A3'].value = float(str(qtd_tot).replace(',', '.'))
        except Exception: ws['A3'].value = qtd_tot
        
        ws['M2'].value = datetime.now().strftime('%d/%m/%Y')
        materiais_com_veio = ["CARVALHO EUROPEU", "OKUME", "LUPA", "PINUS", "ITAPUA", "CARVALHO MEL"]

        row_idx = 6
        for b in sorted(blocos, key=lambda x: 0 if x['tipo'] == 'normal' else 1):
            bloco_e_prensado = (b['tipo'] == 'prensado' or pai_is_prensado)
            
            # Cabeçalho de Bloco Prensado
            if b['tipo'] == 'prensado':
                ws.row_dimensions[row_idx].height = 15.75
                p_d = b['prensado_info']
                ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=13)
                cell_h = ws.cell(row=row_idx, column=2)
                cell_h.value = f"{self.limpar(p_d[1])} - {self.limpar(p_d[3])}"
                cell_h.font = Font(bold=True, size=15)
                cell_h.alignment = align_botao
                row_idx += 1
            
            # Escrita dos Itens
            for item in b['itens']:
                r, is_migrado = row_idx, item.get('is_migrado', False)
                
                if is_migrado:
                    d_l, m_l = item['desc_orig'], item['mat_orig']
                    fita_str, veio_val = str(item['fita_orig']), item['veio_orig']
                    txt_comp = f"{d_l} {m_l}".upper()
                else:
                    desc_f = str(self.limpar(item.get(3, "")))
                    txt_comp = f"{str(item.get(2, ''))} {desc_f} {str(item.get(14, ''))}".upper()
                    d_l, m_l = (desc_f.split(" - ", 1) if " - " in desc_f else ("-", desc_f))
                    if any(m in txt_comp for m in ["KRION", "DURASEIN", "CORIAN", "TS"]): 
                        d_l = str(self.limpar(item.get(14, "")))
                    fita_str = "SEC-LAM" if any(str(self.limpar(item.get(x, ""))) in ['-', '='] for x in [15, 16]) else "SEC"
                    veio_val = None

                is_especial = any(m in txt_comp for m in ["KRION", "DURASEIN", "CORIAN"]) or re.search(r'\bTS\b', txt_comp)
                plus = 5 if (is_especial and not is_migrado) else 0

                try: val_fat = float(item.get('q_unitaria_fatorada', 0))
                except Exception: val_fat = 0.0
                
                # Fórmula de Quantidade
                c_formula = ws.cell(row=r, column=1)
                c_formula.value = f"={val_fat}*A3"
                if val_fat == 0: c_formula.fill = fill_erro
                
                # Dimensões (Comprimento e Largura)
                val_15 = self.limpar(item.get(15, ""))
                v_c_raw = val_15 if val_15 not in ["", "-", "="] else item.get(8, "")
                v_c = self.converter_para_numero(v_c_raw)
                if isinstance(v_c, int): v_c += plus
                ws.cell(row=r, column=2).value = val_15 if val_15 in ["-", "="] else ""
                ws.cell(row=r, column=3).value = v_c
                ws.cell(row=r, column=4).value = "X"

                val_16 = self.limpar(item.get(16, ""))
                v_l_raw = val_16 if val_16 not in ["", "-", "="] else item.get(10, "")
                v_l = self.converter_para_numero(v_l_raw)
                if isinstance(v_l, int): v_l += plus
                ws.cell(row=r, column=5).value = val_16 if val_16 in ["-", "="] else ""
                ws.cell(row=r, column=6).value = v_l
                ws.cell(row=r, column=7).value = "X"

                # Restante dos campos
                ws.cell(row=r, column=8).value = self.converter_para_numero(item.get(12, ""))
                ws.cell(row=r, column=9).value = self.limpar_material_rigoroso(m_l)

                if is_migrado:
                    ws.cell(row=r, column=10).value = veio_val
                else:
                    tem_v = any(m in txt_comp for m in materiais_com_veio)
                    if "KRION" in txt_comp and str(self.converter_para_numero(item.get(12, ""))) == "3": tem_v = True
                    if tem_v: ws.cell(row=r, column=10).value = 1
                
                ws.cell(row=r, column=11).value = fita_str
                ws.cell(row=r, column=12).value = self.converter_para_numero(d_l)
                ws.cell(row=r, column=13).value = self.converter_para_numero(item.get(1, ""))
                
                # Adição visual dos botões de Macro
                if ws.cell(row=r, column=10).value == 1:
                    cel_v = ws.cell(row=r, column=15)
                    cel_v.value = "⇄"; cel_v.fill = fill_botao; cel_v.font = font_veio; cel_v.alignment = align_botao
                
                if not bloco_e_prensado and not is_especial:
                    cel_n = ws.cell(row=r, column=14)
                    cel_n.value = "+5"; cel_n.fill = fill_botao; cel_n.font = font_botao; cel_n.alignment = align_botao
                
                row_idx += 1
        
        self.escrever_seguro(ws, f"A{l_obs}", f"PROJETO DE REFERÊNCIA: {id_proj}", 
                             Alignment(horizontal='left', vertical='center'))
        
        # Salva o arquivo final
        nome_f = re.sub(r'[\\/*?:\u0022<>|]', '', tit).strip()[:120]
        if not nome_f: nome_f = f"PROJETO_SEM_NOME_{cod_p}"
        
        caminho_final = os.path.join(pasta, f"{nome_f}.xlsm")
        tmp_save = caminho_final + ".tmp"
        wb.save(tmp_save)
        if os.path.exists(caminho_final): os.remove(caminho_final)
        os.rename(tmp_save, caminho_final)

    # -------------------------------------------------------------------------
    # CORE: MOTOR DE PROCESSAMENTO
    # -------------------------------------------------------------------------
    def iniciar_processamento(self):
        """Bloqueia a UI e inicia a thread de processamento."""
        self.btn_processar.configure(state="disabled")
        self.progress.grid(row=2, column=0, padx=20, pady=10)
        self.progress.start()
        threading.Thread(target=self.core_processamento, daemon=True).start()

    def core_processamento(self):
        """Lógica principal de interpretação dos dados do Clipboard (PDM)."""
        try:
            is_teste = self.modo_teste_ativo
            df = pd.read_clipboard(sep='\t', header=None, dtype=str).fillna('')
            
            molde = DIRETORIO_SISTEMA / "planilha_molde.xlsm"
            if not molde.exists() and not is_teste: 
                raise Exception(f"Molde não encontrado em: {molde}")

            if df.shape[0] < 2 or df.shape[1] < 6: 
                raise Exception("Dados insuficientes. Copie corretamente do PDM.")
                
            niveis_encontrados = [str(x).count('.') for x in df[0] if re.match(r'^\d+(\.\d+)*$', str(x).strip())]
            if not niveis_encontrados: 
                raise Exception("A estrutura de níveis (coluna 1) não foi identificada.")
                
            id_p = str(df.iloc[1, 1]).strip().upper()
            
            # Definição da Pasta de Destino
            if is_teste: 
                pasta = DESKTOP_PATH / "TESTES_GERADOR" / id_p
            else:
                MARCAS = {"ARO": "Amaro", "BGK": "BurguerKing", "CAM": "Camicado", "CEA": "CeA", "CEN": "Centauro", "ELU": "Elubel", "FAR": "FarmaCopr", "IND": "Indian", "ING": "Ingecon", "MCD": "McDonalds", "PER": "Pernambucanas", "REN": "Renner", "TMS": "Tramontina", "TRA": "Tramontina", "ZAR": "Zara", "ZFR": "Zaffari", "SEP": "Sephora", "PRO": "Prototipo"}
                pasta = DIRETORIO_RAIZ_PLANILHAS / next((v for k,v in MARCAS.items() if k in id_p), "Outros") / id_p
            
            if not os.path.exists(pasta): os.makedirs(pasta)
            
            niv_pai = min(niveis_encontrados)
            if not self.limpar(df.iloc[1, 1]).startswith(('11', '15')): niv_pai += 1

            cache_rede = {} if is_teste else self.mapear_rede_cache()

            # Funções Internas de Filtro
            def f_valido(f):
                c, a, d, mc = str(self.limpar(f.get(1, ""))), str(f.get(2, "")).upper(), str(f.get(3, "")).upper(), str(self.limpar(f.get(14, "")))
                if 'CORTE' in a or any(x in d for x in ["LAMINA MADEIRA", "LAMINA MAD", "LAM MAD", "LAM MADEIRA", "LAMINADO FORM"]) or \
                   any(x in mc for x in ["LAMINA MADEIRA", "LAMINA MAD", "LAM MAD", "LAM MADEIRA", "LAMINADO FORM"]): 
                    return False
                if '*' in a and not c.startswith('11'): return False
                if mc.startswith("92"): 
                    return c.startswith(('11', '15')) if any(m in d for m in ["KRION", "CORIAN", "DURASEIN"]) or re.search(r'\bTS\b', d) else False
                return c.startswith(('11', '15')) and not any(x in mc for x in ["9172", "93"])

            def is_prensado(r):
                return "PRENSADO" in str(r.get(3, "")).upper() or str(self.limpar(r.get(1, ""))) in ["1152032", "1162032"] or "PRLA" in str(r.get(2, "")).upper()

            # Agrupamento de Itens Pai
            cons = {}
            for _, r in df.iterrows():
                nv, cod = self.limpar(r[0]), self.limpar(r[1])
                if cod.startswith(('11', '15')) and nv.count('.') == niv_pai:
                    if nv not in cons: cons[nv] = {'pai': r, 'blocos': [], 'qtd_p_total': 0}
                    cons[nv]['qtd_p_total'] += float(self.converter_para_numero(r[5]) or 0)
            
            arquivos_migrados, projetos_duplicados, arquivos_para_arquivar = [], [], []
            processar_list = []
            
            # Verificação de Duplicidade e Migração
            for nv_p, info in cons.items():
                cod_p = self.limpar(info['pai'][1])
                cam_net = None if is_teste else self.verificar_duplicidade_em_rede(cod_p, cache_rede)

                if cam_net and "PLANOS DE CORTE 2026" not in str(cam_net):
                    itens_mig, a3_mig = self.extrair_dados_migracao(cam_net)
                    if itens_mig:
                        info['blocos'] = [{'tipo': 'normal', 'itens': itens_mig}]
                        info['qtd_p_total'] = a3_mig
                        arquivos_migrados.append(cod_p)
                        processar_list.append((nv_p, info))
                        arquivos_para_arquivar.append(cam_net)
                    else: processar_list.append((nv_p, info))
                elif cam_net: projetos_duplicados.append(cod_p)
                else: processar_list.append((nv_p, info))
            
            # Notificações ao Usuário
            if arquivos_migrados:
                msg = "\n".join(sorted(set(arquivos_migrados)))
                self.after(0, lambda: messagebox.showinfo("Migração Detectada", f"Projetos migrados:\n\n{msg}"))
            
            if projetos_duplicados: 
                msg = ", ".join(sorted(set(projetos_duplicados)))
                self.after(0, lambda: messagebox.showwarning("Repetidos", f"Projetos {msg} já existem e foram ignorados."))

            # Processamento de Estrutura de Filhos
            for nv_p, info in processar_list:
                if not info['blocos']:
                    c_p, desc, p_is_p = self.limpar(info['pai'][1]), df[df[0].str.startswith(nv_p + ".")].copy(), is_prensado(info['pai'])
                    b_roots = {}
                    for _, r in desc.iterrows():
                        nv, cod = self.limpar(r[0]), self.limpar(r[1])
                        if (c_p.startswith('15') and cod.startswith('15') and nv.count('.') > niv_pai) or is_prensado(r):
                            pref = [p for p in b_roots.keys() if nv.startswith(p + ".")]
                            b_roots[nv] = {'tipo': 'prensado', 'prensado_info': r, 'itens': [], 'qf': float(self.converter_para_numero(r[5]) or 1) * (b_roots[max(pref, key=len)]['qf'] if pref else 1.0)}
                    
                    bloco_a = {'tipo': 'normal', 'itens': []}
                    for _, r in desc.iterrows():
                        nv = self.limpar(r[0])
                        pref = [p for p in b_roots.keys() if nv.startswith(p + ".")]
                        parent = b_roots[max(pref, key=len)] if pref else None
                        if not (nv in b_roots) and f_valido(r):
                            ic = r.copy().to_dict()
                            if parent: 
                                ic['q_unitaria_fatorada'] = float(self.converter_para_numero(r[5]) or 0) * parent['qf']
                                parent['itens'].append(ic)
                            elif nv.count('.') == niv_pai + 1: 
                                ic['q_unitaria_fatorada'] = float(self.converter_para_numero(r[5]) or 0)
                                bloco_a['itens'].append(ic)
                    
                    if bloco_a['itens']: info['blocos'].append(bloco_a)
                    for br in b_roots.values(): 
                        if br['itens']: info['blocos'].append(br)
                    
                    if any(len(b['itens']) > 0 for b in info['blocos']):
                        self.gerar_arquivo_excel(info['pai'], info['blocos'], id_p, info['qtd_p_total'], molde, pasta, p_is_p)
                    elif not desc.empty == False and f_valido(info['pai']):
                        ic = info['pai'].copy().to_dict()
                        ic['q_unitaria_fatorada'] = 1.0
                        self.gerar_arquivo_excel(info['pai'], [{'tipo': 'normal', 'itens': [ic]}], id_p, info['qtd_p_total'], molde, pasta, p_is_p)
                else: 
                    self.gerar_arquivo_excel(info['pai'], info['blocos'], id_p, info['qtd_p_total'], molde, pasta, False)

            # Arquivamento de Planilhas Antigas
            if arquivos_para_arquivar and not is_teste:
                if not DIRETORIO_ANTIGOS.exists(): os.makedirs(DIRETORIO_ANTIGOS)
                for arq_antigo in arquivos_para_arquivar:
                    try:
                        shutil.move(arq_antigo, DIRETORIO_ANTIGOS / os.path.basename(arq_antigo))
                    except Exception as e: print(f"[ERRO ARQUIVAR] {e}")

            self.after(0, self.sucesso_final, str(pasta))
        except Exception as e: 
            self.after(0, self.erro_final, str(e))

    # -------------------------------------------------------------------------
    # FINALIZAÇÃO E FEEDBACK
    # -------------------------------------------------------------------------
    def sucesso_final(self, p): 
        self.progress.stop(); self.progress.grid_forget()
        self.btn_processar.configure(state="normal")
        os.startfile(p)
        messagebox.showinfo("Ingecon", "Concluído!")

    def erro_final(self, m): 
        self.progress.stop()
        self.btn_processar.configure(state="normal")
        messagebox.showerror("Erro", str(m))

# =============================================================================
# EXECUÇÃO DO APLICATIVO
# =============================================================================
if __name__ == "__main__": 
    AppIngecon().mainloop()